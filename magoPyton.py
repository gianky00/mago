# -*- coding: utf-8 -*-
import openpyxl
import pyautogui
import time
import pyperclip
import traceback
from datetime import datetime
import sys
import os
import keyboard
import subprocess
import logging
import builtins
import json

# --- CONFIGURAZIONE LOGGING ---
# Configura il logger per scrivere su 'automazione.log'.
# La funzione print() standard verrà usata per l'output nella GUI.
log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
log_file = 'automazione.log'
file_handler = logging.FileHandler(log_file, mode='a', encoding='utf-8')
file_handler.setFormatter(log_formatter)
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
if not logger.hasHandlers():
    logger.addHandler(file_handler)

# --- LIBRERIE OPZIONALI ---
try:
    from PIL import Image
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    print("ATTENZIONE: Librerie 'Pillow' o 'pytesseract' non trovate.")
    OCR_AVAILABLE = False

try:
    import win32com.client as win32
    WIN32_AVAILABLE = True
except ImportError:
    print("ATTENZIONE: Libreria 'pywin32' non trovata.")
    WIN32_AVAILABLE = False

# --- FUNZIONI DI CONFIGURAZIONE ---
def load_config():
    """Carica la configurazione dal file config.json."""
    try:
        with open('config.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print("ERRORE CRITICO: File 'config.json' non trovato.")
        return None
    except json.JSONDecodeError:
        print("ERRORE CRITICO: Il file 'config.json' non è formattato correttamente.")
        return None

# --- FUNZIONI DI AUTOMAZIONE ---
def restart_script(config):
    """Riavvia lo script corrente."""
    TASTO_PER_RIAVVIARE = config['tasti_rapidi']['riavvia']
    current_pid = os.getpid()
    print(f"\n--- RICHIESTA DI RIAVVIAO (TASTO '{TASTO_PER_RIAVVIARE}' PREMUTO) (PID: {current_pid}) ---")
    try:
        keyboard.remove_all_hotkeys()
    except Exception:
        pass
    python_executable = sys.executable
    script_to_run = os.path.abspath(sys.argv[0])
    subprocess.Popen([python_executable, script_to_run] + sys.argv[1:])
    sys.exit(0)

def force_excel_recalculation(filepath):
    if not WIN32_AVAILABLE:
        print("ERRORE: Impossibile forzare il ricalcolo, pywin32 non disponibile.")
        return False
    if not os.path.exists(filepath):
        print(f"ERRORE: File non trovato per il ricalcolo: {filepath}")
        return False

    print(f"\nTentativo di forzare il ricalcolo di Excel per: {os.path.basename(filepath)}...")
    excel_instance = None
    try:
        excel_instance = win32.Dispatch('Excel.Application')
        excel_instance.Visible = False
        excel_instance.DisplayAlerts = False
        excel_instance.EnableEvents = False
        wb = excel_instance.Workbooks.Open(filepath)
        wb.Application.CalculateFullRebuild()
        wb.Save()
        wb.Close(SaveChanges=False)
        excel_instance.Quit()
        print("Ricalcolo e salvataggio completati.")
        return True
    except Exception as e:
        print(f"\nERRORE durante il ricalcolo forzato di Excel: {e}")
        traceback.print_exc()
        return False
    finally:
        if excel_instance:
            excel_instance.Quit()

def mostra_dialogo_input(titolo, messaggio):
    """Funzione helper per mostrare un dialogo di input e ottenere il risultato."""
    import tkinter as tk
    from tkinter import simpledialog
    root = tk.Tk()
    root.withdraw()
    risultato = simpledialog.askstring(titolo, messaggio, parent=root)
    root.destroy()
    return risultato

def controlla_regione_per_testo(config, odc_cfg, regione, testo_da_cercare, nome_log, click_coords=None):
    """
    Funzione OCR generica per controllare una regione alla ricerca di un testo.
    Se il testo viene trovato, opzionalmente esegue un click.
    Restituisce True se il testo viene trovato, altrimenti False.
    """
    if not OCR_AVAILABLE:
        print(f"ATTENZIONE: OCR non disponibile, impossibile controllare '{nome_log}'.")
        return False
    # Step 1: Validazione della regione
    try:
        # Converte in interi e si assicura che siano 4 valori
        x, y, w, h = map(int, regione)
        # Se la larghezza o l'altezza sono 0, la regione non è valida per uno screenshot
        if w <= 0 or h <= 0:
            return False
        regione_valida = (x, y, w, h)
    except (ValueError, TypeError):
        # Se la regione non è una lista/tupla di 4 numeri, non è valida
        return False

    # Step 2: Tentativo di cattura e OCR
    try:
        pytesseract.pytesseract.tesseract_cmd = config['file_e_fogli_excel']['impostazioni_file']['path_tesseract_cmd']
        time.sleep(odc_cfg.get('pausa_attesa_popup', 0.5))
        screenshot = pyautogui.screenshot(region=regione_valida)
        testo_estratto = pytesseract.image_to_string(screenshot, lang='ita')

        # Step 3: Controllo del testo e azione (con log solo in caso di successo)
        if testo_da_cercare.lower() in testo_estratto.lower():
            print(f"  --> POPUP RILEVATO: '{nome_log}'. Azione in corso...")
            if click_coords:
                pyautogui.click(click_coords)
            time.sleep(0.5)
            return True

        return False # Nessun log se il testo non viene trovato

    except Exception as e:
        # Logga solo errori di esecuzione reali, non il "tile cannot extend"
        # che viene prevenuto dalla validazione all'inizio.
        print(f"\nERRORE durante l'analisi OCR del popup '{nome_log}': {e}")
        return False

def esegui_procedura_registrazione_odc(config, odc_cfg, gui_cfg, valore_odc_fallito, dati_riga_completa):
    """Esegue la procedura di registrazione di un nuovo ODC."""
    print("\n--- INIZIO PROCEDURA REGISTRAZIONE ODC ---")
    try:
        pyautogui.click(odc_cfg['coordinate_popup_tasto_no']); time.sleep(1.0)
        pyautogui.click(odc_cfg['coordinate_barra_preferiti']); time.sleep(1.0)
        pyautogui.click(odc_cfg['coordinate_commesse_variante']); time.sleep(1.5)
        pyautogui.click(gui_cfg['coordinate_apertura_nuovo_foglio']); time.sleep(1.0)

        # Inserimento codice e descrizione
        pyautogui.click(odc_cfg['coordinate_casella_commessa'])
        pyautogui.write(str(valore_odc_fallito), interval=0.05)
        pyautogui.press('tab'); time.sleep(0.5)

        descrizione = dati_riga_completa.get('T', '') # Assumendo che 'T' sia la colonna descrizione
        pyautogui.click(odc_cfg['coordinate_casella_descrizione'])
        if descrizione:
            pyperclip.copy(str(descrizione))
            pyautogui.hotkey('ctrl', 'v')
        pyautogui.press('tab'); time.sleep(0.5)

        # Gestione popup intermedi durante la registrazione
        # Nota: questi popup sono gestiti con la logica "trova testo e clicca"
        controlla_regione_per_testo(config, odc_cfg, odc_cfg['regione_popup_attenzione_sal'], odc_cfg['testo_popup_attenzione_sal'], "Attenzione_SAL", click_coords=odc_cfg['coordinate_click_popup_attenzione_sal'])
        controlla_regione_per_testo(config, odc_cfg, odc_cfg['regione_popup_codice_iva'], odc_cfg['testo_popup_codice_iva'], "Codice_IVA", click_coords=odc_cfg['coordinate_click_popup_codice_iva'])

        # Completamento campi rimanenti
        pyautogui.click(odc_cfg['coordinate_casella_commessa_padre']); pyautogui.write("23/134", interval=0.05); pyautogui.press('tab'); time.sleep(0.5)
        pyautogui.press('tab'); time.sleep(0.3); pyautogui.press('tab'); time.sleep(0.3)
        pyautogui.click(odc_cfg['coordinate_casella_ns_rif']); pyautogui.write("1765", interval=0.05); pyautogui.press('tab'); time.sleep(0.5)
        pyautogui.click(odc_cfg['coordinate_apri_elenco_gruppo']); time.sleep(1.0); pyautogui.click(odc_cfg['coordinate_casella_gruppo_generico']); time.sleep(0.5)
        pyautogui.click(odc_cfg['coordinate_casella_resp_commessa']); pyautogui.write("1765", interval=0.05); pyautogui.press('tab'); time.sleep(0.5)
        pyautogui.click(odc_cfg['coordinate_casella_resp_cantiere']); pyautogui.write("1765", interval=0.05); pyautogui.press('tab'); time.sleep(0.5)
        pyautogui.click(odc_cfg['coordinate_barra_causali_depositi']); time.sleep(1.0); pyautogui.click(odc_cfg['coordinate_apri_elenco_deposito']); time.sleep(1.0); pyautogui.click(odc_cfg['coordinate_casella_deposito']); time.sleep(0.5)

        # Salvataggio e chiusura
        pyautogui.click(gui_cfg['coordinate_salvataggio_foglio']); time.sleep(2.0)
        pyautogui.click(odc_cfg['coordinate_chiudi_rapportino_commessa']); time.sleep(1.5)

        print("--- PROCEDURA REGISTRAZIONE ODC COMPLETATA ---")
        return True
    except Exception as e:
        print(f"\nERRORE durante la procedura ODC: {e}")
        traceback.print_exc()
        return False

def esegui_copia_da_buffer_e_verifica(config, valore_da_copiare_str, cella_excel_id_sorgente):
    timing = config['timing_e_ritardi']
    pulizia = config['pulizia_appunti']
    for _ in range(timing['tentativi_max_copia_excel']):
        try:
            pyperclip.copy('')
            time.sleep(pulizia['pausa_1'])
            pyperclip.copy('---PULIZIA---')
            time.sleep(pulizia['pausa_2'])
            pyperclip.copy('')
            time.sleep(pulizia['pausa_3'])
            time.sleep(timing['ritardo_prima_copia_verifica'])
            pyperclip.copy(valore_da_copiare_str)
            time.sleep(timing['ritardo_dopo_copia_excel'])
            if pyperclip.paste() == valore_da_copiare_str:
                pyperclip.copy(valore_da_copiare_str)
                time.sleep(pulizia['riconferma_copia_pausa'])
                return True
        except Exception as e:
            print(f"ERRORE copia buffer: {e}")
        time.sleep(timing['ritardo_tra_tentativi_copia'])
    print(f"FALLIMENTO COPIA DA BUFFER per {cella_excel_id_sorgente}.")
    return False

def esegui_incolla_singolo_click(config, valore_incollato_str, target_coords, select_all=False):
    timing = config['timing_e_ritardi']
    pyperclip.copy(valore_incollato_str)
    time.sleep(timing['ritardo_prima_copia_finale'])
    pyautogui.moveTo(target_coords[0], target_coords[1], duration=timing['ritardo_movimento_mouse'])
    time.sleep(timing['ritardo_dopo_movimento'])
    pyautogui.click()
    time.sleep(timing['ritardo_click_singolo'])
    if select_all:
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(timing['ritardo_dopo_select_all'])
    time.sleep(timing['ritardo_prima_incolla'])
    pyautogui.hotkey('ctrl', 'v')

def esegui_incolla_e_tab(config, valore_incollato_str, target_coords, cella_excel_id):
    timing = config['timing_e_ritardi']
    pyperclip.copy(valore_incollato_str)
    time.sleep(timing['ritardo_prima_copia_finale'])
    pyautogui.moveTo(target_coords[0], target_coords[1], duration=timing['ritardo_movimento_mouse'])
    time.sleep(timing['ritardo_dopo_movimento'])
    pyautogui.click()
    time.sleep(timing['ritardo_click_singolo'])
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(timing['ritardo_dopo_select_all'])
    time.sleep(timing['ritardo_prima_incolla'])
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('tab')
    time.sleep(timing['ritardo_dopo_tab'])


def aggiorna_matricola_excel(config, vecchia_matricola, nuova_matricola):
    """
    Cerca una matricola nel foglio 'MATRICOLE' e la aggiorna.
    """
    file_cfg = config['file_e_fogli_excel']['impostazioni_file']
    NOME_FILE_EXCEL = os.path.abspath(file_cfg['percorso_file_excel'])
    NOME_FOGLIO_MATRICOLE = "MATRICOLE"

    print(f"\n--- AGGIORNAMENTO MATRICOLA IN EXCEL ---")
    print(f"  File: {os.path.basename(NOME_FILE_EXCEL)}")
    print(f"  Cerco '{vecchia_matricola}' per sostituirla con '{nuova_matricola}'...")

    wb = None
    try:
        wb = openpyxl.load_workbook(NOME_FILE_EXCEL, keep_vba=True)
        if NOME_FOGLIO_MATRICOLE not in wb.sheetnames:
            print(f"  ERRORE: Foglio '{NOME_FOGLIO_MATRICOLE}' non trovato nel file Excel.")
            return False

        sheet = wb[NOME_FOGLIO_MATRICOLE]

        matricola_trovata = False
        # Itera sulla colonna A a partire dalla riga 2
        for riga in range(2, sheet.max_row + 1):
            cella = sheet[f'A{riga}']
            if str(cella.value).strip() == str(vecchia_matricola).strip():
                print(f"  --> Trovata corrispondenza alla riga {riga}. Aggiorno il valore...")
                cella.value = nuova_matricola
                matricola_trovata = True
                break

        if not matricola_trovata:
            print(f"  ATTENZIONE: Matricola '{vecchia_matricola}' non trovata nel foglio '{NOME_FOGLIO_MATRICOLE}'. Nessuna modifica apportata.")
            return False

        wb.save(NOME_FILE_EXCEL)
        print("  Salvataggio del file Excel completato.")
        return True

    except Exception as e:
        print(f"  ERRORE CRITICO durante l'aggiornamento della matricola in Excel: {e}")
        traceback.print_exc()
        return False
    finally:
        if wb:
            wb.close()


# --- FUNZIONE PRINCIPALE DI AUTOMAZIONE ---
def run_automation(config):
    # --- Setup delle configurazioni ---
    file_cfg = config['file_e_fogli_excel']['impostazioni_file']
    param_cfg = config['file_e_fogli_excel']['mappature_colonne_foglio_avanzamento']
    gui_cfg = config['coordinate_e_dati']['gui']
    timing_cfg = config['timing_e_ritardi']
    tasti_cfg = config['tasti_rapidi']

    # --- Caricamento del profilo attivo ---
    try:
        profili_cfg = config['file_e_fogli_excel']['mappatura_colonne_profili']
        profilo_attivo_nome = profili_cfg['profilo_attivo']
        profilo_attivo_dati = profili_cfg['profili'][profilo_attivo_nome]
        col_mapping = profilo_attivo_dati['mappature']
        odc_cfg = profilo_attivo_dati['impostazioni_odc'] # Carica la configurazione ODC specifica del profilo
        print(f"INFO: Caricato profilo di automazione: '{profilo_attivo_nome}'")
    except KeyError as e:
        print(f"ERRORE CRITICO: Chiave di configurazione mancante nel profilo '{profilo_attivo_nome}': {e}")
        return

    NOME_FILE_EXCEL = os.path.abspath(file_cfg['percorso_file_excel'])

    print(f"--- AVVIO AUTOMAZIONE (PID: {os.getpid()}) ---")
    try:
        keyboard.add_hotkey(tasti_cfg['riavvia'], lambda: restart_script(config), suppress=False)
        print(f"INFO: Premi '{tasti_cfg['riavvia']}' per riavviare.")
    except Exception as e:
        print(f"ATTENZIONE: Impossibile impostare hotkey: {e}")

    if file_cfg['forzare_ricalcolo_excel']:
        if WIN32_AVAILABLE:
            if not force_excel_recalculation(NOME_FILE_EXCEL):
                print("ERRORE CRITICO: Ricalcolo Excel fallito.")
                return
        else:
            print("ATTENZIONE: Ricalcolo abilitato ma pywin32 non disponibile.")
    else:
        print("INFO: Ricalcolo forzato di Excel disabilitato.")

    print(f"\nFASE 1: Lettura dati da '{os.path.basename(NOME_FILE_EXCEL)}'...")

    colonne_da_leggere = [item['colonna_excel'] for item in col_mapping]
    STATO_DA_CERCARE = "DA COMPLETARE"

    task_da_eseguire = None
    try:
        wb_leggi = openpyxl.load_workbook(NOME_FILE_EXCEL, data_only=True, keep_vba=True)
        sheet_parametri = wb_leggi[file_cfg['nome_foglio_avanzamento']]
        sheet_dati = wb_leggi[file_cfg['nome_foglio_dati']]

        # cantiere = param_cfg.get('ns_riferimento', '') # This seems to be unused now, keeping it commented for safety

        col_stato = sheet_parametri[param_cfg['colonna_stato']]
        for cella_stato in col_stato:
            stato_raw = cella_stato.value
            stato_pulito = str(stato_raw).strip().upper() if stato_raw is not None else ""

            if stato_pulito == STATO_DA_CERCARE:
                r = cella_stato.row
                print(f"  --> TASK TROVATO ALLA RIGA {r}!")
                r_inizio = sheet_parametri[f"{param_cfg['colonna_riga_inizio']}{r}"].value
                r_fine = sheet_parametri[f"{param_cfg['colonna_riga_fine']}{r}"].value
                data_raw = sheet_parametri[f"{param_cfg['colonna_data_rapporto']}{r}"].value
                data_rapporto = data_raw.strftime('%d/%m/%Y') if isinstance(data_raw, datetime) else str(data_raw)

                dati_buffer = []
                for r_dati in range(r_inizio, r_fine + 1):
                    dati_riga = {col: sheet_dati[f"{col}{r_dati}"].value for col in colonne_da_leggere}
                    if any(dati_riga.values()):
                        dati_buffer.append({'riga_excel_num': r_dati, 'dati_colonne': dati_riga})

                if dati_buffer:
                    task_da_eseguire = {
                        'riga_param_task': r,
                        'data_rapporto': data_rapporto,
                        'dati_buffer': dati_buffer
                    }
                    print(f"  Trovato task valido (riga {r}), bufferizzate {len(dati_buffer)} righe.")
                    break
        wb_leggi.close()
    except Exception as e:
        print(f"ERRORE CRITICO FASE 1: {e}"); traceback.print_exc()
        return

    if not task_da_eseguire:
        print("\n--- NESSUN TASK VALIDO TROVATO ---")
        return

    print(f"\nHai 2 secondi per preparare le finestre... (Premi '{tasti_cfg['riavvia']}' per riavviare)")
    for i in range(2, 0, -1): print(f"{i}...", end=" ", flush=True); time.sleep(1)
    print("Via!")

    errore_gui = False
    try:
        # Azioni preliminari
        pyautogui.click(gui_cfg['coordinate_apertura_nuovo_foglio']); time.sleep(timing_cfg['pausa_tra_azioni_preliminari_finali'])
        if task_da_eseguire['data_rapporto']:
            esegui_incolla_singolo_click(config, task_da_eseguire['data_rapporto'], gui_cfg['coordinate_data_rapporto'], select_all=True)
            time.sleep(timing_cfg['pausa_tra_azioni_preliminari_finali'])

        if gui_cfg.get('dato_ns_riferimento'):
            esegui_incolla_singolo_click(config, str(gui_cfg['dato_ns_riferimento']), gui_cfg['coordinate_cantiere'])
            pyautogui.press('tab'); time.sleep(timing_cfg['ritardo_dopo_tab'])

        # --- Ciclo sui dati ---
        righe_processate_blocco = 0
        try:
            commessa_mapping = next(item for item in col_mapping if item["nome"] == "Commessa")
            matricola_mapping = next(item for item in col_mapping if item["nome"] == "Matricola")
            descrizione_mapping = next((item for item in col_mapping if item["nome"] == "Descrizione Attività"), None)
        except StopIteration as e:
            print(f"ERRORE CRITICO: Mappatura per 'Commessa' o 'Matricola' non trovata nel profilo '{profilo_attivo_nome}'. Impossibile continuare.")
            return

        for i, riga_obj in enumerate(task_da_eseguire['dati_buffer']):
            y_target = timing_cfg['y_iniziale_remoto'] + (righe_processate_blocco * timing_cfg['incremento_y_remoto'])
            print(f"\nProcessando riga {i+1}/{len(task_da_eseguire['dati_buffer'])} (Excel: {riga_obj['riga_excel_num']})...")

            # --- Gestione dei campi principali ---
            riga_da_saltare = False
            for mapping_item in col_mapping:
                if riga_da_saltare: break
                if mapping_item['nome'] == "Descrizione Attività": continue

                col_lettera = mapping_item['colonna_excel']
                target_x = mapping_item['target_x_remoto']
                val = riga_obj['dati_colonne'].get(col_lettera)

                if val is not None:
                    val_str = str(val)
                    cella_id = f"{file_cfg['nome_foglio_dati']}!{col_lettera}{riga_obj['riga_excel_num']}"
                    if not esegui_copia_da_buffer_e_verifica(config, val_str, cella_id):
                        raise Exception(f"Errore fatale di copia dal buffer per {cella_id}")

                    if target_x > 0:
                        esegui_incolla_e_tab(config, val_str, (target_x, y_target), cella_id)

                        # --- LOGICA POST-INSERIMENTO ---
                        # Caso 1: Inserimento Commessa
                        if mapping_item['nome'] == commessa_mapping['nome']:
                            if controlla_regione_per_testo(config, odc_cfg, odc_cfg['regione_popup_odc_gia_registrato_specifico'], "esisten", "ODC_Gia_Registrato_Specifico"):
                                print("  SUCCESS: ODC già registrato (caso specifico). Salto alla riga successiva.")
                                pyautogui.click(odc_cfg['coordinate_chiudi_commessa_variante']); time.sleep(0.5)
                                pyautogui.click(odc_cfg['coordinate_non_salvare_commessa_variante']); time.sleep(0.5)
                                riga_da_saltare = True; break
                            if 'regione_popup_odc_gia_registrato' in odc_cfg and controlla_regione_per_testo(config, odc_cfg, odc_cfg['regione_popup_odc_gia_registrato'], odc_cfg.get('testo_popup_odc_gia_registrato', 'già registrata'), "ODC_Gia_Registrato"):
                                print("  --> Rilevato popup 'ODC Già Registrato'. Chiudo le finestre e salto la riga.")
                                pyautogui.click(odc_cfg['coordinate_chiudi_commessa_variante']); time.sleep(0.5)
                                pyautogui.click(odc_cfg['coordinate_non_salvare_commessa_variante']); time.sleep(0.5)
                                riga_da_saltare = True; break
                            if controlla_regione_per_testo(config, odc_cfg, odc_cfg['regione_screenshot_popup_generico'], odc_cfg['testo_da_cercare_popup_generico'], "Non_Trovato_Commessa"):
                                print("  INFO: Popup 'non trovato' per Commessa. Avvio registrazione nuovo ODC.")
                                if not esegui_procedura_registrazione_odc(config, odc_cfg, gui_cfg, val_str, riga_obj['dati_colonne']):
                                    raise Exception("Fallimento procedura registrazione ODC.")
                                print("  Reinserisco la commessa per continuare il flusso normale.")
                                esegui_incolla_e_tab(config, val_str, (target_x, y_target), cella_id)

                        # Caso 2: Inserimento Matricola
                        elif mapping_item['nome'] == matricola_mapping['nome']:
                            matricola_corretta = False
                            tentativi_disattivo = 0
                            TENTATIVI_MAX_DISATTIVO = 2 # Trigger dialog on the second failure
                            while not matricola_corretta:
                                # Check 0: Matricola "disattivo"
                                if 'regione_popup_matricola_disattivo' in odc_cfg and controlla_regione_per_testo(config, odc_cfg, odc_cfg['regione_popup_matricola_disattivo'], odc_cfg.get('testo_popup_matricola_disattivo', 'disattivo'), "Matricola_Disattivo"):
                                    tentativi_disattivo += 1
                                    print(f"  --> Rilevato popup 'disattivo' per matricola (Tentativo {tentativi_disattivo}/{TENTATIVI_MAX_DISATTIVO}).")
                                    pyautogui.click(odc_cfg['coordinate_popup_tasto_no']); time.sleep(1.0)

                                    if tentativi_disattivo >= TENTATIVI_MAX_DISATTIVO:
                                        # Second attempt failed, show dialog to user
                                        col_nome = odc_cfg.get('colonna_nome_dipendente', 'F')
                                        nome_dipendente = riga_obj['dati_colonne'].get(col_nome, '[NOME NON TROVATO]')
                                        messaggio = f"La matricola {val_str} per {nome_dipendente} è ancora disattivata.\n\nInserisci la nuova matricola corretta:"
                                        nuova_matricola = mostra_dialogo_input("Matricola Disattivata", messaggio)

                                        if nuova_matricola and nuova_matricola.strip():
                                            vecchia_matricola = val_str
                                            val_str = nuova_matricola.strip() # Update with new value
                                            print(f"  INPUT: Nuova matricola '{val_str}' fornita. Verrà aggiornata in Excel e riprovata.")

                                            # Call the new function to update the Excel file
                                            aggiorna_matricola_excel(config, vecchia_matricola, val_str)

                                            tentativi_disattivo = 0 # Reset counter for the new matricola
                                            esegui_incolla_e_tab(config, val_str, (target_x, y_target), "matricola_sostitutiva_disattivo")
                                            continue
                                        else:
                                            print("  WARNING: L'utente ha annullato l'inserimento della nuova matricola. Riga saltata.")
                                            riga_da_saltare = True
                                            break # Exit while loop
                                    else: # First attempt failed, just retry
                                        print("  ...Riprovo a inserire la stessa matricola.")
                                        esegui_incolla_e_tab(config, val_str, (target_x, y_target), cella_id)
                                        continue

                                # Check 1: Matricola disabilitata
                                if 'regione_popup_matricola_disabilitata' in odc_cfg and controlla_regione_per_testo(config, odc_cfg, odc_cfg['regione_popup_matricola_disabilitata'], "disabilitata", "Matricola_Disabilitata"):
                                    pyautogui.press('escape') # Chiude il popup di matricola disabilitata
                                    nuova_matricola = mostra_dialogo_input("Matricola Disabilitata", f"La matricola {val_str} è disabilitata.\nInserisci la nuova matricola corretta:")
                                    if nuova_matricola:
                                        val_str = nuova_matricola # Sovrascrive il valore per il resto del ciclo
                                        print(f"  INPUT: Nuova matricola '{val_str}' fornita. Riprovo l'inserimento.")
                                        esegui_incolla_e_tab(config, val_str, (target_x, y_target), "matricola_sostitutiva")
                                        continue # Ritorna all'inizio del while per riverificare la nuova matricola
                                    else:
                                        print("  WARNING: Nessuna nuova matricola fornita. Riga saltata.")
                                        riga_da_saltare = True; break

                                # Check 2: Matricola non trovata (1° tentativo)
                                if controlla_regione_per_testo(config, odc_cfg, odc_cfg['regione_screenshot_popup_generico'], odc_cfg['testo_da_cercare_popup_generico'], "Non_Trovato_Matricola_1"):
                                    print("  WARNING: Popup 'non trovato' per Matricola. Riprovo (1/2)...")
                                    esegui_incolla_e_tab(config, val_str, (target_x, y_target), cella_id) # Riprova

                                    # Check 2.1: Matricola non trovata (2° tentativo)
                                    if controlla_regione_per_testo(config, odc_cfg, odc_cfg['regione_screenshot_popup_generico'], odc_cfg['testo_da_cercare_popup_generico'], "Non_Trovato_Matricola_2"):
                                        print("  ERROR: Popup 'non trovato' anche al secondo tentativo.")
                                        pyautogui.click(odc_cfg['coordinate_popup_tasto_no']) # Clicca "No"
                                        col_nome = odc_cfg.get('colonna_nome_dipendente', 'F')
                                        nome_dipendente = riga_obj['dati_colonne'].get(col_nome, '[NOME NON TROVATO]')
                                        nuova_matricola_utente = mostra_dialogo_input("Errore Matricola", f"Matricola errata per {nome_dipendente}.\nFornisci la Matricola corretta:")
                                        if nuova_matricola_utente:
                                            val_str = nuova_matricola_utente
                                            print(f"  INPUT: Nuova matricola '{val_str}' fornita. Riprovo l'inserimento.")
                                            esegui_incolla_e_tab(config, val_str, (target_x, y_target), "matricola_corretta_utente")
                                            continue # Ritorna all'inizio del while per riverificare
                                        else:
                                            print("  WARNING: Riga saltata su richiesta utente dopo errore matricola.")
                                            riga_da_saltare = True; break
                                    else:
                                        matricola_corretta = True # OK al secondo tentativo
                                else:
                                    matricola_corretta = True # OK al primo tentativo

                            if riga_da_saltare: break
                else:
                    if target_x > 0:
                        pyautogui.press('tab'); time.sleep(timing_cfg['ritardo_dopo_tab'])

                # Inserimento Descrizione Attività (se presente e se la matricola è stata appena inserita)
                if not riga_da_saltare and mapping_item['nome'] == matricola_mapping['nome'] and descrizione_mapping:
                    val_desc = riga_obj['dati_colonne'].get(descrizione_mapping['colonna_excel'])
                    if val_desc is not None:
                        val_str_desc = str(val_desc)
                        cella_id_desc = f"{file_cfg['nome_foglio_dati']}!{descrizione_mapping['colonna_excel']}{riga_obj['riga_excel_num']}"
                        if not esegui_copia_da_buffer_e_verifica(config, val_str_desc, cella_id_desc):
                            raise Exception(f"Errore copia GUI per {cella_id_desc}")
                        if descrizione_mapping['target_x_remoto'] > 0:
                            esegui_incolla_e_tab(config, val_str_desc, (descrizione_mapping['target_x_remoto'], y_target), cella_id_desc)

            if not riga_da_saltare:
                righe_processate_blocco += 1
                if righe_processate_blocco >= timing_cfg['max_righe_per_blocco_tab']:
                    print("\n  Raggiunto limite blocco, premo Tab per cambiare pagina...")
                    for _ in range(timing_cfg['pressioni_tasto_tab']): pyautogui.press('tab', interval=timing_cfg['intervallo_pressioni_tab'])
                    righe_processate_blocco = 0
                    time.sleep(timing_cfg['pausa_dopo_blocco_tab'])
        print("\n--- FINE AZIONI GUI ---")

    except Exception as e:
        print(f"\nERRORE CRITICO DURANTE AZIONI GUI: {e}"); traceback.print_exc()
        errore_gui = True

    if not errore_gui:
        print(f"\nPremi '{tasti_cfg['prosegui']}' per salvare 'OK' in Excel ed eseguire le azioni finali...")
        keyboard.wait(tasti_cfg['prosegui'])

        file_salvato = False
        try:
            wb_scrivi = openpyxl.load_workbook(NOME_FILE_EXCEL, keep_vba=True)
            sheet_dati_scrivi = wb_scrivi[file_cfg['nome_foglio_dati']]
            colonna_ok_dati = config.get('mapping_colonne', {}).get('colonna_ok_dati', 'A') # Fallback to 'A'
            for riga_obj in task_da_eseguire['dati_buffer']:
                sheet_dati_scrivi[f"{colonna_ok_dati}{riga_obj['riga_excel_num']}"] = "OK"

            for _ in range(timing_cfg['tentativi_max_salvataggio_excel']):
                try:
                    wb_scrivi.save(NOME_FILE_EXCEL)
                    file_salvato = True
                    print("File Excel salvato con successo.")
                    break
                except Exception as e_save:
                    print(f"Tentativo di salvataggio fallito: {e_save}")
                    time.sleep(timing_cfg['ritardo_tra_tentativi_salvataggio'])
            wb_scrivi.close()
        except Exception as e:
            print(f"ERRORE CRITICO SCRITTURA/SALVATAGGIO EXCEL: {e}"); traceback.print_exc()

        if file_salvato:
            print("Esecuzione azioni finali GUI...")
            pyautogui.click(gui_cfg['coordinate_salvataggio_foglio']); time.sleep(timing_cfg['pausa_tra_azioni_preliminari_finali'])
            pyautogui.click(gui_cfg['coordinate_elimina_flag_stampa']); time.sleep(timing_cfg['pausa_tra_azioni_preliminari_finali'])
            pyautogui.click(gui_cfg['coordinate_pulsante_ok']); time.sleep(timing_cfg['pausa_tra_azioni_preliminari_finali'])
            print("--- AUTOMAZIONE COMPLETATA CON SUCCESSO ---")
        else:
            print("--- AUTOMAZIONE INTERROTTA: SALVATAGGIO EXCEL FALLITO ---")
    else:
        print("--- AUTOMAZIONE INTERROTTA A CAUSA DI ERRORI GUI ---")

    print(f"\nPremi '{tasti_cfg['riavvia']}' per un nuovo ciclo o chiudi la finestra.")
    try:
        while True: time.sleep(1)
    except KeyboardInterrupt:
        print("\nScript terminato.")

# This block is intentionally left empty so this file can be used as a library
if __name__ == "__main__":
    pass